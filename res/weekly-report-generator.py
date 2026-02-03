import sys
import shutil
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

weekday_kor = {
    "Mon": "월", "Tue": "화", "Wed": "수", "Thu": "목",
    "Fri": "금", "Sat": "토", "Sun": "일"
}

def friday_of_week(dt: datetime) -> datetime:
    """주어진 날짜가 속한 주의 금요일을 반환"""
    return dt - timedelta(days=dt.weekday()) + timedelta(days=4)

def get_submit_date() -> datetime:
    """사용자로부터 날짜를 입력받거나 기본값을 반환"""
    user_input = input("작성 날짜를 입력(YYYYMMDD): ").strip()
    if user_input == "":
        return datetime.today()
    else:
        return datetime.strptime(user_input, "%Y%m%d")

def auto_fit_rows(file_path: str, sheet_name: str):
    """Excel을 실제로 열어서 행 높이 자동 맞춤 실행"""
    try:
        import win32com.client
    except ImportError:
        print("경고: pywin32가 설치되지 않아 행 높이 자동 맞춤을 건너뜁니다.")
        print("설치: pip install pywin32")
        return
    
    print("Excel을 열어 행 높이를 자동 맞춤 중...")
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # 백그라운드 실행
        excel.DisplayAlerts = False  # 경고 메시지 숨김
        
        # 절대 경로로 파일 열기
        workbook = excel.Workbooks.Open(file_path)
        worksheet = workbook.Worksheets(sheet_name)
        
        # 전체 행에 대해 자동 맞춤
        worksheet.Rows.AutoFit()
        
        workbook.Save()
        workbook.Close()
        print("행 높이 자동 맞춤 완료!")
    except Exception as e:
        print(f"행 높이 자동 맞춤 중 오류 발생: {e}")
    finally:
        if excel:
            excel.Quit()

def main():
    # 경로 설정 (스크립트가 res 폴더에 있으므로 상위 폴더가 작업 루트)
    script_dir = os.path.dirname(os.path.realpath(__file__))
    base_path = os.path.dirname(script_dir)  # res의 상위 폴더
    print("작업경로: %s" % base_path)

    # 날짜 결정
    submit = get_submit_date()
    
    # 연도 추출 및 연도 폴더 생성
    year = submit.strftime("%Y")
    year_folder = os.path.join(base_path, year)
    
    if not os.path.exists(year_folder):
        os.makedirs(year_folder)
        print(f"{year}년 폴더를 생성했습니다.")

    # ISO 주차
    WW = submit.strftime("%V")
    WW_int = int(WW)
    WW_old_int = WW_int - 1
    WW_old = f"{WW_old_int:02d}"

    # 리소스 및 템플릿 로드
    res_path = os.path.join(base_path, "res")
    
    # res 폴더 존재 확인
    if not os.path.exists(res_path):
        print("[오류] res 폴더를 찾을 수 없습니다.")
        print(f"경로: {res_path}")
        print("\n프로젝트 구조를 확인해주세요.")
        input("종료하려면 Enter 키를 누르세요...")
        sys.exit(1)
    
    file_res = os.listdir(res_path)
    xlsx_templates = [f for f in file_res if f.endswith('.xlsx') and not f.startswith('~$')]
    
    # 템플릿 파일 존재 확인
    if not xlsx_templates:
        print("[오류] res 폴더에 템플릿 파일(.xlsx)이 없습니다.")
        print(f"경로: {res_path}")
        print("\nWW00으로 시작하는 Excel 템플릿 파일을 res 폴더에 넣어주세요.")
        print("예시: WW00_부서명_업무보고서_이름.xlsx")
        input("종료하려면 Enter 키를 누르세요...")
        sys.exit(1)
    
    name_form = xlsx_templates[0]
    print(f"템플릿 파일: {name_form}")
    
    # 연도 폴더의 기존 파일 목록
    xlsx_files = [f for f in os.listdir(year_folder) if f.endswith('.xlsx')]
    
    new_name = name_form[0:2] + WW + name_form[4:]
    last_name = name_form[0:2] + WW_old + name_form[4:]
    
    new_file_path = os.path.join(year_folder, new_name)
    last_file_path = os.path.join(year_folder, last_name)

    # ---- 중복 파일 검사: 인터랙티브 ----
    if new_name in xlsx_files:
        check = 0
        while check not in ["Y", "y"]:
            check = input("이미 해당 주의 파일이 존재합니다. 덮어쓰시겠습니까?(Y/N): ")
            if check in ["N", "n"]:
                sys.exit(0)

    # 새로운 문서 생성
    shutil.copyfile(os.path.join(res_path, name_form), new_file_path)
    WB = load_workbook(new_file_path)
    WS = WB.worksheets[0]  # 첫 번째 시트 자동 선택
    make_date = submit.strftime("%Y-%m-%d")
    WS['A4'].value = "Date : %s" % make_date

    # 상반기(월~금)
    day_insert = submit - timedelta(days=submit.weekday())
    for i in range(0, 5):
        mth = int(day_insert.strftime("%m"))
        day = int(day_insert.strftime("%d"))
        wkd = weekday_kor[day_insert.strftime("%a")]
        WS.cell(8 + i, 1).value = "%d/%d(%s)" % (mth, day, wkd)
        day_insert = day_insert + timedelta(days=1)

    # 하반기(월~금, 전주 로직 유지: 토/일 건너뛰고 5일)
    day_insert = day_insert + timedelta(days=2)
    for i in range(0, 5):
        mth = int(day_insert.strftime("%m"))
        day = int(day_insert.strftime("%d"))
        wkd = weekday_kor[day_insert.strftime("%a")]
        WS.cell(16 + i, 1).value = "%d/%d(%s)" % (mth, day, wkd)
        day_insert = day_insert + timedelta(days=1)

    # 시트명 변경 및 저장
    WS.title = "WW%s" % WW

    # 데이터 유효성 추가
    dv = DataValidation(type="list", formula1="{0}!$B$2:$B$5".format(quote_sheetname('Sheet1')), showErrorMessage=True, allow_blank=True, showDropDown=False)
    dv.add('B8:B12')
    dv.add('B16:B20')
    WS.add_data_validation(dv)
    
    WB.save(new_file_path)

    # ---- 이전 주 내용 가져오기: 인터랙티브 ----
    if last_name in xlsx_files:
        check = 0
        while check not in ["Y", "y"]:
            check = input("이전 주의 파일이 존재합니다. 기록된 내용을 가져오시겠습니까?(Y/N): ")
            if check in ["N", "n"]:
                sys.exit(0)
        
        WB_old = load_workbook(last_file_path)
        WS_old = WB_old["WW%s" % WW_old]
        cp = 4
        while WS_old.cell(cp - 2, 1).value != "차주일정":
            cp += 1
        
        for i in range(0, 5):
            for j in range(0, 3):
                WS.cell(8 + i, 3 + j).value = WS_old.cell(cp + i, 3 + j).value
            for k in range(0, 2):
                WS.cell(8 + i, 7 + k).value = WS_old.cell(cp + i, 6 + k).value
        
        for i in range(0, 3):
            for j in range(0, 3):
                WS.cell(25 + i, 1 + j).value = WS_old.cell(cp + 9 + i, 1 + j).value
            for k in range(0, 2):
                WS.cell(25 + i, 5 + k).value = WS_old.cell(cp + 9 + i, 5 + k).value
        
        WB.save(new_file_path)

    # ---- pywin32를 사용한 행 높이 자동 맞춤 ----
    auto_fit_rows(os.path.abspath(new_file_path), f"WW{WW}")

    print(f"파일 생성 완료: {year}\\{new_name}")
    print("Job Done!")
    input() # 종료 전 대기

if __name__ == "__main__":
    main()
