import sys
import shutil
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

weekday_kor = {
    "Mon": "월", "Tue": "화", "Wed": "수", "Thu": "목",
    "Fri": "금", "Sat": "토", "Sun": "일"
}

def friday_of_week(dt: datetime) -> datetime:
    """주어진 날짜가 속한 주의 금요일을 반환"""
    return dt - timedelta(days=dt.weekday()) + timedelta(days=4)

def parse_submit_from_argv() -> datetime:
    """명령행 인수가 있으면 YYYYMMDD로 해석, 없으면 실행 주 금요일 사용"""
    if len(sys.argv) >= 2:
        arg = sys.argv[1].strip()
        if len(arg) == 8 and arg.isdigit():
            return datetime.strptime(arg, "%Y%m%d")
        else:
            raise ValueError(f"잘못된 날짜 인수입니다(YYYYMMDD 필요): '{arg}'")
    # 인수가 없으면 자동: 실행 주 금요일
    return friday_of_week(datetime.today())

def auto_fit_rows(file_path: str, sheet_name: str):
    """Excel을 실제로 열어서 행 높이 자동 맞춤 실행"""
    try:
        import win32com.client
    except ImportError:
        print("경고: pywin32가 설치되지 않아 행 높이 자동 맞춤을 건너뜁니다.")
        print("설치: pip install pywin32")
        return
    
    print("Excel을 열어 행 높이를 자동 맞춤 중...")
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # 백그라운드 실행
        excel.DisplayAlerts = False  # 경고 메시지 숨김
        
        # 절대 경로로 파일 열기
        workbook = excel.Workbooks.Open(file_path)
        worksheet = workbook.Worksheets(sheet_name)
        
        # 전체 행에 대해 자동 맞춤
        worksheet.Rows.AutoFit()
        
        workbook.Save()
        workbook.Close()
        print("행 높이 자동 맞춤 완료!")
    except Exception as e:
        print(f"행 높이 자동 맞춤 중 오류 발생: {e}")
    finally:
        if excel:
            excel.Quit()

def main():
    # 경로 설정 (스크립트가 res 폴더에 있으므로 상위 폴더가 작업 루트)
    script_dir = os.path.dirname(os.path.realpath(__file__))
    base_path = os.path.dirname(script_dir)  # res의 상위 폴더
    print("작업경로: %s" % base_path)

    # 날짜 결정: 인수 또는 자동 금요일
    submit = parse_submit_from_argv()
    
    # 연도 추출 및 연도 폴더 생성
    year = submit.strftime("%Y")
    year_folder = os.path.join(base_path, year)
    
    if not os.path.exists(year_folder):
        os.makedirs(year_folder)
        print(f"{year}년 폴더를 생성했습니다.")

    # ISO 주차
    WW = submit.strftime("%V")
    WW_int = int(WW)
    WW_old_int = WW_int - 1
    WW_old = f"{WW_old_int:02d}"

    # 리소스 및 템플릿 로드
    res_path = os.path.join(base_path, "res")
    file_res = os.listdir(res_path)
    name_form = [f for f in file_res if f.endswith('.xlsx')][0]

    # 연도 폴더의 기존 파일 목록
    xlsx_files = [f for f in os.listdir(year_folder) if f.endswith('.xlsx')]

    new_name = name_form[0:2] + WW + name_form[4:]
    last_name = name_form[0:2] + WW_old + name_form[4:]
    
    new_file_path = os.path.join(year_folder, new_name)
    last_file_path = os.path.join(year_folder, last_name)

    # ---- 중복 파일 검사: 기본값 N(건너뛰기) ----
    if new_name in xlsx_files:
        print(f"이미 해당 주의 파일이 존재합니다. 생성 작업을 건너뜁니다: {year}\\{new_name}")
        # 정상 종료(자동화 파이프라인에서 실패로 간주되지 않도록 0 종료)
        sys.exit(0)

    # 새로운 문서 생성
    shutil.copyfile(os.path.join(res_path, name_form), new_file_path)
    WB = load_workbook(new_file_path)
    WS = WB.worksheets[0]  # 첫 번째 시트 자동 선택
    make_date = submit.strftime("%Y-%m-%d")
    WS['A4'].value = "Date : %s" % make_date

    # 상반기(월~금)
    day_insert = submit - timedelta(days=submit.weekday())
    for i in range(0, 5):
        mth = int(day_insert.strftime("%m"))
        day = int(day_insert.strftime("%d"))
        wkd = weekday_kor[day_insert.strftime("%a")]
        WS.cell(8 + i, 1).value = "%d/%d(%s)" % (mth, day, wkd)
        day_insert = day_insert + timedelta(days=1)

    # 하반기(월~금, 다음 주 월~금이 아니라 기존 로직 유지: 토/일 건너뛴 뒤 5일)
    day_insert = day_insert + timedelta(days=2)
    for i in range(0, 5):
        mth = int(day_insert.strftime("%m"))
        day = int(day_insert.strftime("%d"))
        wkd = weekday_kor[day_insert.strftime("%a")]
        WS.cell(16 + i, 1).value = "%d/%d(%s)" % (mth, day, wkd)
        day_insert = day_insert + timedelta(days=1)

    # 시트명 변경 및 저장
    WS.title = "WW%s" % WW

    # 데이터 유효성 추가
    dv = DataValidation(type="list", formula1="{0}!$B$2:$B$5".format(quote_sheetname('Sheet1')), showErrorMessage=True, allow_blank=True, showDropDown=False)
    dv.add('B8:B12')
    dv.add('B16:B20')
    WS.add_data_validation(dv)

    WB.save(new_file_path)

    # ---- 이전 주 내용 가져오기: 기본값 Y(자동 가져오기) ----
    if last_name in xlsx_files:
        try:
            WB_old = load_workbook(last_file_path)
            WS_old = WB_old["WW%s" % WW_old]
            cp = 4
            while WS_old.cell(cp - 2, 1).value != "차주일정":
                cp += 1

            for i in range(0, 5):
                for j in range(0, 3):
                    WS.cell(8 + i, 3 + j).value = WS_old.cell(cp + i, 3 + j).value
                for k in range(0, 2):
                    WS.cell(8 + i, 7 + k).value = WS_old.cell(cp + i, 6 + k).value

            for i in range(0, 3):
                for j in range(0, 3):
                    WS.cell(25 + i, 1 + j).value = WS_old.cell(cp + 9 + i, 1 + j).value
                for k in range(0, 2):
                    WS.cell(25 + i, 5 + k).value = WS_old.cell(cp + 9 + i, 5 + k).value

            WB.save(new_file_path)
            print(f"이전 주({WW_old}) 내용 자동 반영 완료.")
        except Exception as e:
            # 이전 주 파일/시트 구조가 예상과 다를 경우에도 자동화가 중단되지 않도록 스킵
            print(f"이전 주 내용 가져오기 중 오류가 발생했지만 스킵합니다: {e}")

    # ---- pywin32를 사용한 행 높이 자동 맞춤 ----
    auto_fit_rows(os.path.abspath(new_file_path), f"WW{WW}")

    print(f"파일 생성 완료: {year}\\{new_name}")
    print("Job Done!")

if __name__ == "__main__":
    main()
